"""
择时因子评价框架 — 主程序
=========================

使用方式
--------
1. 将 price_df.csv 和 signal.csv 放入 data_ini/ 目录
2. 编辑 config.yaml 中的资产代码、信号名称及各项参数
3. 运行：python main.py

输出
----
- workspace/plots/eval/      : 因子评估图（每个信号8子图 + 多因子对比图）
- workspace/plots/backtest/  : 回测图（每个信号3张图）
- workspace/timing_report.xlsx : 完整 Excel 报告（6个分析Sheet）

数据格式要求
-----------
price_df.csv
    支持以下列命名格式（会自动匹配）：
    1. 列名即为资产代码，如 "000300.SH"（CLOSE价格直接在该列）
    2. 列名为 "{资产代码}_{字段}"，如 "000300.SH_CLOSE"
    3. 多级索引列，如 ("000300.SH", "CLOSE")
    4. 单资产文件，列名直接为 "CLOSE" 或 "close"
    索引须为日期格式（可被 pd.to_datetime 解析）。

signal.csv
    每列为一个信号，列名即信号名称。
    索引须为日期格式，值为连续因子数值。
"""

from __future__ import annotations

import os
import sys
import warnings
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ── 路径配置：确保能导入 timing_framework ──────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))
warnings.filterwarnings("ignore")

# ── Matplotlib 后端与字体（必须在其他 import 之前设置）──────────────────────
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "WenQuanYi Micro Hei",
                                    "PingFang SC", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False

import numpy as np
import pandas as pd

# ── 第三方库（带友好报错）──────────────────────────────────────────────────
try:
    import yaml
except ImportError:
    raise ImportError("请安装 PyYAML：pip install pyyaml")

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
        numbers as xl_numbers,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
except ImportError:
    raise ImportError("请安装 openpyxl：pip install openpyxl")

# ── 内部模块 ──────────────────────────────────────────────────────────────
from timing_framework import FactorPreprocessor, TimingFactorEvaluator
from backtest.backtester import Backtester, BacktestResult


# ══════════════════════════════════════════════════════════════════════════════
# 工具函数
# ══════════════════════════════════════════════════════════════════════════════

def load_config(config_path: str | Path = "config.yaml") -> dict:
    """加载 YAML 配置文件，返回配置字典。"""
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    return cfg


def load_data(data_dir: str | Path = "data_ini",
             price_file: str = "price_df.csv",
             signal_file: str = "signal.csv") -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    从 data_ini/ 目录加载价格和信号 CSV 文件。将第一列日期列转换为索引，并按照日期格式处理和排序

    文件名通过 config.yaml 的 data.price_file / data.signal_file 指定。

    返回
    ----
    (price_df, signal_df)，索引均为 DatetimeIndex。
    """
    data_dir = Path(data_dir)
    price_path  = data_dir / price_file
    signal_path = data_dir / signal_file

    if not price_path.exists():
        raise FileNotFoundError(f"价格文件不存在：{price_path.resolve()}")
    if not signal_path.exists():
        raise FileNotFoundError(f"信号文件不存在：{signal_path.resolve()}")

    price_df  = pd.read_csv(price_path,  index_col=0, parse_dates=True)
    signal_df = pd.read_csv(signal_path, index_col=0, parse_dates=True)

    price_df.index  = pd.to_datetime(price_df.index)
    signal_df.index = pd.to_datetime(signal_df.index)

    price_df  = price_df.sort_index()
    signal_df = signal_df.sort_index()

    return price_df, signal_df


def extract_price_series(price_df: pd.DataFrame,
                          asset_code: str,
                          price_field: str = "CLOSE") -> pd.Series:
    """
    从多资产价格 DataFrame 中智能提取指定资产的价格序列。
    虽然建议在数据表字段按照‘datetime-code-CLOSE’名称严格设置，但这里依旧给予了几项可能的匹配策略。

    """
    #策略main: 如果结构式 datetime-code-CLOSE，直接提取
    code_columns_name = ['code','asset_code','wind_code','sec_code']
    for code_col in code_columns_name:
        if code_col in price_df.columns:
            print(f"匹配资产代码列名：{code_col}")
            price_df = price_df[price_df[code_col] == asset_code]
            return price_df[price_field].dropna().rename("price")


    # 策略1：如果结构是 datetime-code1-code2-code3
    if asset_code in price_df.columns:
        print(f"直接匹配资产代码列：{asset_code}")
        return price_df[asset_code].dropna().rename("price")

    # 策略2：如果结构是 datetime-code1_CLOSE-code2-code3
    col2 = f"{asset_code}_{price_field}"
    if col2 in price_df.columns:
        print(f"匹配资产代码和字段名：{col2}")
        return price_df[col2].dropna().rename("price")

    # 策略3：单资产，结构为 datetime-CLOSE
    pf_lower = price_field.lower()
    exact_matches = [c for c in price_df.columns
                     if isinstance(c, str) and c.lower() == pf_lower]
    if len(exact_matches) == 1:
        print(f"直接匹配字段名（大小写敏感）：{exact_matches[0]}")
        return price_df[exact_matches[0]].dropna().rename("price")

    # 策略4：MultiIndex 列
    if isinstance(price_df.columns, pd.MultiIndex):
        if (asset_code, price_field) in price_df.columns:
            print(f"匹配 MultiIndex 资产代码和字段名：{asset_code, price_field}")
            return price_df[(asset_code, price_field)].dropna().rename("price")
        if (asset_code, price_field.lower()) in price_df.columns:
            print(f"匹配 MultiIndex 资产代码和字段名（大小写敏感）：{asset_code, price_field.lower()}")
            return price_df[(asset_code, price_field.lower())].dropna().rename("price")

    # 策略5：大小写不敏感模糊匹配（含资产代码和字段名）
    for col in price_df.columns:
        if isinstance(col, str):
            cl = col.lower()
            if asset_code.lower() in cl and pf_lower in cl:
                print(f"大小写不敏感模糊匹配：{col}")
                return price_df[col].dropna().rename("price")

    # 均未匹配，给出明确的错误信息
    raise ValueError(
        f"无法在 price_df 中找到资产 '{asset_code}' 的 '{price_field}' 价格列。\n"
        f"请检查 config.yaml 的 asset.code 和 asset.price_field 设置。\n"
        f"price_df 现有列（前20个）：{list(price_df.columns[:20])}"
    )


def select_signals(signal_df: pd.DataFrame, cfg_signals: dict) -> pd.DataFrame:
    """
    根据配置选择信号列。

    若 use_all=true，返回全部列；否则返回 selected 列表中的列。
    """
    if cfg_signals.get("use_all", False):
        return signal_df

    selected = cfg_signals.get("selected", [])
    if not selected:
        raise ValueError("config.yaml 中 signals.use_all=false 但 signals.selected 为空，"
                         "请至少指定一个信号名称。")

    missing = [s for s in selected if s not in signal_df.columns]
    if missing:
        raise ValueError(
            f"以下信号在 signal.csv 中不存在：{missing}\n"
            f"signal.csv 现有列：{list(signal_df.columns)}"
        )
    return signal_df[selected]


def _prepare_eval_data(
    raw_signal: pd.Series,
    prices: pd.Series,
    rolling_window: int,
) -> Tuple[pd.Series, pd.Series, pd.Series, int, str, int]:
    """
    根据信号的实际观测频率，自动准备评估专用数据集。

    问题背景
    --------
    若把低频信号（如周度/月度）前向填充到日度后再计算 IC，
    滚动窗口内会存在大量重复值（同一信号值与多个日度收益配对），
    导致有效自由度被严重高估、T 统计量失真。

    解决方案：评估与回测分离
    -------------------------
    - **评估**：信号保持自然频率，收益率同步降频到信号观测日之间的
      "期间收益"（P[t]/P[t-1]-1，t 为相邻信号日），IC 无重复计数。
    - **回测**：信号前向填充到日度，使仓位在信号更新前保持不变，
      符合实际交易逻辑（由调用方处理）。

    频率检测
    --------
    用价格日历（交易日）统计相邻信号观测日的中位间隔，自动判断频率。
    不依赖 config，任意不规则频率均可处理。

    参数
    ----
    raw_signal     : 原始信号序列（含 NaN，dropna 后为实际观测日）
    prices         : 日度收盘价序列（交易日日历）
    rolling_window : config 中指定的滚动窗口（交易日数，默认 252）

    返回
    ----
    eval_signal     : 信号观测日上的因子序列（dropna 后）
    eval_prices     : 与信号观测日对齐的收盘价
    eval_returns    : 相邻信号观测日之间的期间收益率
    adjusted_window : 按频率等比缩放后的滚动窗口（信号观测数）
    freq_label      : 可读频率标签（"日度"/"周度"/"月度" 等）
    median_gap      : 中位交易日间隔（=1 即为日度）
    """
    signal_obs = raw_signal.dropna()
    if len(signal_obs) < 5:
        raise ValueError(
            f"信号有效观测点不足（{len(signal_obs)} 条），无法评估。"
        )

    # ── 非交易日对齐：将信号日期映射到最近的上一个交易日 ────────────
    # 规则：
    #   - 若目标交易日本身已有信号值 → 直接丢弃该非交易日信号（不覆盖）
    #   - 若目标交易日为空           → 映射过去
    price_idx = prices.index
    not_in_calendar = ~signal_obs.index.isin(price_idx)
    if not_in_calendar.any():
        n_total = int(not_in_calendar.sum())

        trading_signals     = signal_obs[~not_in_calendar]   # 已在交易日上的信号
        non_trading_signals = signal_obs[not_in_calendar]    # 需要处理的非交易日信号

        # 找每个非交易日对应的最近上一交易日
        snap_pos    = price_idx.searchsorted(non_trading_signals.index, side="right") - 1
        snap_pos    = np.clip(snap_pos, 0, len(price_idx) - 1)
        snapped_dates = price_idx[snap_pos]

        # 已被占用的交易日（初始为所有已有信号的交易日）
        occupied = set(trading_signals.index)

        n_mapped, n_dropped = 0, 0
        mapped_items: dict = {}

        for snapped_date, value in zip(snapped_dates, non_trading_signals.values):
            if snapped_date in occupied:
                # 目标交易日已有值 → 丢弃
                n_dropped += 1
            else:
                # 目标交易日为空 → 映射
                mapped_items[snapped_date] = value
                occupied.add(snapped_date)   # 防止后续信号再次映射到同一天
                n_mapped += 1

        if mapped_items:
            extra = pd.Series(mapped_items, name=signal_obs.name)
            signal_obs = pd.concat([trading_signals, extra]).sort_index()
        else:
            signal_obs = trading_signals

        print(f"  [日期对齐] {n_total} 个信号不在交易日历中："
              f"{n_mapped} 个映射到最近上一交易日，{n_dropped} 个因目标日已有值而丢弃")

    # ── 检测频率：用价格日历计量相邻信号日之间的交易日数 ─────────────
    pos_in_price = price_idx.searchsorted(signal_obs.index, side="left")
    # 截断越界（信号日晚于价格最后一天的情况）
    pos_in_price = np.clip(pos_in_price, 0, len(price_idx) - 1)
    gaps = np.diff(pos_in_price)
    median_gap = int(np.median(gaps)) if len(gaps) > 0 else 1
    median_gap = max(1, median_gap)

    # ── 频率标签 ────────────────────────────────────────────────────────
    if   median_gap <= 2:  freq_label = "日度"
    elif median_gap <= 4:  freq_label = "2-3日"
    elif median_gap <= 7:  freq_label = "周度"
    elif median_gap <= 12: freq_label = "半月度"
    elif median_gap <= 25: freq_label = "月度"
    elif median_gap <= 70: freq_label = "季度"
    else:                  freq_label = f"约{median_gap}交易日/次"

    # ── 对齐价格到信号观测日（取最近交易日收盘价）─────────────────────
    eval_prices = prices.reindex(signal_obs.index, method="ffill")

    # ── 期间收益率：相邻信号观测日之间的累积收益 ─────────────────────
    # eval_returns[t_k] = P[t_k]/P[t_{k-1}] - 1
    # evaluator 内部会再 shift(-n) 以获取 n 期前瞻收益
    eval_returns = eval_prices.pct_change()

    # ── 日度信号：直接使用日度收益，滚动窗口不需调整 ─────────────────
    if median_gap <= 2:
        daily_returns = prices.pct_change().reindex(signal_obs.index)
        return signal_obs, eval_prices, daily_returns, rolling_window, freq_label, median_gap

    # ── 非日度信号：等比缩放滚动窗口（保持约 1 年时间跨度）───────────
    adjusted_window = max(10, round(rolling_window / median_gap))
    return signal_obs, eval_prices, eval_returns, adjusted_window, freq_label, median_gap


# ══════════════════════════════════════════════════════════════════════════════
# Excel 报告生成器
# ══════════════════════════════════════════════════════════════════════════════

class ExcelReporter:
    """
    将择时因子评估与回测结果写入格式化 Excel 报告。

    Sheet 布局
    ----------
    1. 综合概览     : 全信号汇总对比表
    2. IC相关性分析  : 多周期 IC / ICIR 详细数据
    3. 信号检验结果  : 三种方法的胜率与盈亏指标
    4. 回归分析     : OLS 系数、T值、P值、R²
    5. 稳健性检验   : 样本内外、市场状态分析
    6. 回测绩效汇总  : 策略 vs 基准绩效指标对比
    """

    # 颜色方案
    _C_HEADER   = "1565C0"  # 深蓝（表头背景）
    _C_SUBHEAD  = "E3F2FD"  # 浅蓝（次级表头）
    _C_POS      = "C8E6C9"  # 浅绿（正值高亮）
    _C_NEG      = "FFCDD2"  # 浅红（负值高亮）
    _C_ROW_ALT  = "F8F9FA"  # 交替行底色
    _C_TITLE    = "0D47A1"  # 标题字体颜色

    def __init__(self) -> None:
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # 删除默认空表

    # ── 样式辅助 ─────────────────────────────────────────────────────────────

    def _hfill(self, hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def _border(self, style: str = "thin") -> Border:
        s = Side(style=style)
        return Border(left=s, right=s, top=s, bottom=s)

    def _header_font(self, bold: bool = True, color: str = "FFFFFF",
                     size: int = 10) -> Font:
        return Font(bold=bold, color=color, name="微软雅黑", size=size)

    def _data_font(self, bold: bool = False, size: int = 10) -> Font:
        return Font(bold=bold, name="微软雅黑", size=size)

    def _center(self) -> Alignment:
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _write_header_row(self, ws, row: int, cols: List[str],
                           col_start: int = 1) -> None:
        """写入带样式的表头行。"""
        for j, text in enumerate(cols, start=col_start):
            c = ws.cell(row=row, column=j, value=text)
            c.fill      = self._hfill(self._C_HEADER)
            c.font      = self._header_font()
            c.alignment = self._center()
            c.border    = self._border()

    def _write_data_row(self, ws, row: int, values: list,
                        col_start: int = 1, alt: bool = False) -> None:
        """写入数据行，可选交替底色。"""
        fill = self._hfill(self._C_ROW_ALT) if alt else None
        for j, v in enumerate(values, start=col_start):
            c = ws.cell(row=row, column=j, value=v)
            c.font      = self._data_font()
            c.alignment = self._center()
            c.border    = self._border("thin")
            if fill:
                c.fill = fill

    def _auto_width(self, ws, min_w: int = 8, max_w: int = 25) -> None:
        """自动调整列宽（基于内容长度）。"""
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = (
                min(max_w, max(min_w, max_len + 2))
            )

    def _pct(self, v) -> str:
        """将浮点数格式化为百分比字符串，None/NaN 显示为 N/A。"""
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "N/A"
        return f"{v:.2%}"

    def _f4(self, v) -> str:
        """浮点数 4 位小数，None/NaN 显示 N/A。"""
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "N/A"
        return f"{v:.4f}"

    def _color_cell(self, cell, value, threshold: float = 0.0,
                    reverse: bool = False) -> None:
        """根据正负值为单元格着色（正→绿，负→红；reverse=True 时相反）。"""
        if value is None or (isinstance(value, float) and np.isnan(value)):
            return
        is_good = value > threshold if not reverse else value < threshold
        cell.fill = self._hfill(self._C_POS if is_good else self._C_NEG)

    # ── Sheet 1：综合概览 ──────────────────────────────────────────────────

    def _sheet_overview(self,
                         results: List[Tuple],
                         asset_name: str,
                         cfg: dict) -> None:
        ws = self.wb.create_sheet("综合概览")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A3"

        # 大标题（合并范围随列数同步：14列 → 16列）
        ws.merge_cells("A1:P1")
        title_cell = ws["A1"]
        title_cell.value = f"择时因子评价报告  |  资产：{asset_name}"
        title_cell.font      = Font(bold=True, size=14, color=self._C_TITLE, name="微软雅黑")
        title_cell.alignment = self._center()
        title_cell.fill      = self._hfill("EBF5FB")
        ws.row_dimensions[1].height = 28

        headers = [
            "信号名称", "预测周期", "IC方法",
            "IC均值", "ICIR", "IC正值占比",
            "胜率(阈值法)", "胜率(均线法)", "胜率(极值法)",
            "β系数", "β显著性", "R²",
            "样本外稳健", "IC衰减幅度",
            "综合评分", "评级",
        ]
        self._write_header_row(ws, row=2, cols=headers)
        ws.row_dimensions[2].height = 22

        for idx, (sig_name, ev, bt_res) in enumerate(results):
            row = idx + 3
            alt = (idx % 2 == 1)

            # 取各指标
            ic1 = (ev._ic_results or {}).get(ev.forward_period)
            sig1 = (ev._signal_results or {}).get("threshold")
            sig2 = (ev._signal_results or {}).get("moving_average")
            sig3 = (ev._signal_results or {}).get("percentile")
            reg  = ev._reg_result
            rob  = ev._robustness_result
            sc   = ev.score()

            ic_mean   = ic1.ic_mean           if ic1  else None
            icir      = ic1.icir              if ic1  else None
            ic_pos_r  = ic1.ic_positive_ratio if ic1  else None
            wr1 = sig1.full.overall_win_rate if sig1 else None
            wr2 = sig2.full.overall_win_rate if sig2 else None
            wr3 = sig3.full.overall_win_rate if sig3 else None
            beta      = reg.beta       if reg  else None
            beta_sig  = "显著" if (reg and reg.is_significant) else "不显著"
            r2        = reg.r_squared  if reg  else None
            is_robust = "稳健" if (rob and rob.is_robust) else "不稳健"
            ic_deg    = rob.ic_degradation if rob else None

            values = [
                sig_name,
                f"{ev.forward_period}期",
                ev.ic_method.upper(),
                self._f4(ic_mean),  self._f4(icir),   self._pct(ic_pos_r),
                self._pct(wr1),     self._pct(wr2),   self._pct(wr3),
                self._f4(beta),     beta_sig,          self._f4(r2),
                is_robust,          self._pct(ic_deg),
                f"{sc.composite_score:.4f}",
                sc.grade,
            ]
            self._write_data_row(ws, row=row, values=values, alt=alt)

            # 着色：IC/ICIR/胜率/评分（列索引随新增列右移2位）
            cells = [ws.cell(row, j+1) for j in range(len(values))]
            for ci, val in [(3, ic_mean), (4, icir)]:
                self._color_cell(cells[ci], val if val is not None else 0)
            for ci, val in [(6, wr1), (7, wr2), (8, wr3)]:
                self._color_cell(cells[ci], val - 0.5 if val is not None else None)
            # 评级列着色（第15列，索引15）
            grade_colors = {"A": "B2EBF2", "B": "C8E6C9", "C": "FFF9C4",
                            "D": "FFE0B2", "F": "FFCDD2"}
            cells[15].fill = self._hfill(grade_colors.get(sc.grade, "FFFFFF"))

        self._auto_width(ws)

    # ── Sheet 2：IC相关性分析 ───────────────────────────────────────────────

    def _sheet_ic(self, results: List[Tuple]) -> None:
        ws = self.wb.create_sheet("IC相关性分析")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        headers = [
            "信号名称", "预测周期", "IC方法",
            "IC均值", "IC标准差", "ICIR",
            "IC>0占比", "T统计量", "P值", "是否显著",
        ]
        self._write_header_row(ws, row=1, cols=headers)

        row = 2
        for sig_name, ev, _ in results:
            for period, ic_res in sorted((ev._ic_results or {}).items()):
                alt = (row % 2 == 0)
                values = [
                    sig_name, f"{period}日", ic_res.method.upper(),
                    self._f4(ic_res.ic_mean),
                    self._f4(ic_res.ic_std),
                    self._f4(ic_res.icir),
                    self._pct(ic_res.ic_positive_ratio),
                    self._f4(ic_res.t_statistic),
                    self._f4(ic_res.p_value),
                    "✓ 显著" if ic_res.is_significant else "✗ 不显著",
                ]
                self._write_data_row(ws, row=row, values=values, alt=alt)
                # ICIR 着色
                c = ws.cell(row, 6)
                self._color_cell(c, ic_res.icir if not np.isnan(ic_res.icir) else 0)
                row += 1

        self._auto_width(ws)

    # ── Sheet 3：信号检验结果 ───────────────────────────────────────────────

    def _sheet_signal(self, results: List[Tuple]) -> None:
        ws = self.wb.create_sheet("信号检验结果")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        headers = [
            "信号名称", "检验方法",
            "多头样本数", "多头胜率", "多头均收益", "多头盈亏比",
            "空头样本数", "空头胜率", "空头均收益", "空头盈亏比",
            "综合胜率", "多空收益价差", "T统计量", "P值", "是否显著",
            "切分日期",
            "IS多头胜率", "IS多头盈亏比",
            "OOS多头胜率", "OOS多头盈亏比",
        ]
        self._write_header_row(ws, row=1, cols=headers)

        method_label = {
            "threshold":     "阈值法",
            "moving_average":"均线法",
            "percentile":    "极值法",
        }

        row = 2
        for sig_name, ev, _ in results:
            for method, res in (ev._signal_results or {}).items():
                alt = (row % 2 == 0)
                full = res.full
                # IS/OOS 数据
                split_d = str(res.split_date)[:10] if res.split_date else "N/A"
                is_lwr  = self._pct(res.insample.long_win_rate)  if res.insample  else "N/A"
                is_lpl  = self._f4(res.insample.long_pl_ratio)   if res.insample  else "N/A"
                oos_lwr = self._pct(res.outsample.long_win_rate) if res.outsample else "N/A"
                oos_lpl = self._f4(res.outsample.long_pl_ratio)  if res.outsample else "N/A"
                values = [
                    sig_name, method_label.get(method, method),
                    full.n_long,
                    self._pct(full.long_win_rate),
                    self._f4(full.long_avg_return),
                    self._f4(full.long_pl_ratio),
                    full.n_short,
                    self._pct(full.short_win_rate),
                    self._f4(full.short_avg_return),
                    self._f4(getattr(full, "short_pl_ratio", None)),
                    self._pct(full.overall_win_rate),
                    self._f4(full.long_short_return_spread),
                    self._f4(full.t_statistic),
                    self._f4(full.p_value),
                    "✓ 显著" if full.is_significant else "✗ 不显著",
                    split_d,
                    is_lwr, is_lpl,
                    oos_lwr, oos_lpl,
                ]
                self._write_data_row(ws, row=row, values=values, alt=alt)
                # 胜率着色
                self._color_cell(ws.cell(row, 4), full.long_win_rate - 0.5)
                self._color_cell(ws.cell(row, 11), full.overall_win_rate - 0.5)
                row += 1

        self._auto_width(ws)

    # ── Sheet 4：回归分析 ──────────────────────────────────────────────────

    def _sheet_regression(self, results: List[Tuple]) -> None:
        ws = self.wb.create_sheet("回归分析")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        headers = [
            "信号名称",
            "Alpha", "Beta",
            "Alpha-T值", "Beta-T值",
            "Alpha-P值", "Beta-P值",
            "R²", "调整R²",
            "F统计量", "F-P值",
            "样本量", "β是否显著",
        ]
        self._write_header_row(ws, row=1, cols=headers)

        for idx, (sig_name, ev, _) in enumerate(results):
            reg  = ev._reg_result
            alt  = (idx % 2 == 1)
            if reg is None:
                values = [sig_name] + ["N/A"] * 12
            else:
                values = [
                    sig_name,
                    self._f4(reg.alpha),      self._f4(reg.beta),
                    self._f4(reg.alpha_tstat), self._f4(reg.beta_tstat),
                    self._f4(reg.alpha_pvalue),self._f4(reg.beta_pvalue),
                    self._f4(reg.r_squared),   self._f4(reg.adj_r_squared),
                    self._f4(reg.f_statistic),
                    self._f4(getattr(reg, "f_pvalue", None)),
                    reg.n_observations,
                    "✓ 显著" if reg.is_significant else "✗ 不显著",
                ]
            self._write_data_row(ws, row=idx + 2, values=values, alt=alt)
            if reg:
                self._color_cell(ws.cell(idx + 2, 3), reg.beta)   # beta 着色
                self._color_cell(ws.cell(idx + 2, 8),
                                  reg.r_squared - 0.01)             # R² 着色

        self._auto_width(ws)

    # ── Sheet 5：稳健性检验 ────────────────────────────────────────────────

    def _sheet_robustness(self, results: List[Tuple]) -> None:
        ws = self.wb.create_sheet("稳健性检验")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        headers = [
            "信号名称",
            "样本分割日期",
            "样本内IC均值", "样本内ICIR",
            "样本外IC均值", "样本外ICIR",
            "IC衰减幅度",   "是否稳健",
            "牛市IC均值",   "牛市ICIR",
            "熊市IC均值",   "熊市ICIR",
            "震荡市IC均值", "震荡市ICIR",
        ]
        self._write_header_row(ws, row=1, cols=headers)

        for idx, (sig_name, ev, _) in enumerate(results):
            rob   = ev._robustness_result
            reg   = ev._regime_results or {}
            alt   = (idx % 2 == 1)

            def _regime_val(key: str, attr: str):
                r = reg.get(key)
                val = getattr(r, attr, None) if r else None
                return self._f4(val)

            if rob is None:
                values = [sig_name] + ["N/A"] * 13
            else:
                values = [
                    sig_name,
                    str(rob.split_date)[:10] if rob.split_date else "N/A",
                    self._f4(rob.insample_ic),   self._f4(rob.insample_icir),
                    self._f4(rob.outsample_ic),  self._f4(rob.outsample_icir),
                    self._pct(rob.ic_degradation),
                    "✓ 稳健" if rob.is_robust else "✗ 不稳健",
                    _regime_val("bull",     "ic_mean"),
                    _regime_val("bull",     "icir"),
                    _regime_val("bear",     "ic_mean"),
                    _regime_val("bear",     "icir"),
                    _regime_val("sideways", "ic_mean"),
                    _regime_val("sideways", "icir"),
                ]
            self._write_data_row(ws, row=idx + 2, values=values, alt=alt)
            if rob:
                # 稳健性列着色
                self._color_cell(ws.cell(idx + 2, 8),
                                  1.0 if rob.is_robust else -1.0)

        self._auto_width(ws)

    # ── Sheet 6：回测绩效汇总 ─────────────────────────────────────────────

    def _sheet_backtest(self, results: List[Tuple], benchmark_ratio: float) -> None:
        ws = self.wb.create_sheet("回测绩效汇总")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        headers = [
            "信号名称",
            "总收益率(策略)", "年化收益(策略)", "年化波动(策略)",
            "Sharpe(策略)",  "最大回撤(策略)", "Calmar(策略)",
            "日胜率(策略)",  "交易次数",
            "交易胜率",      "持有时间占比",   "赔率",          "超额胜率",
            "总收益率(基准)", "年化收益(基准)", "Sharpe(基准)",
            "最大回撤(基准)",
            "超额年化收益",   "超额Sharpe",
        ]
        self._write_header_row(ws, row=1, cols=headers)

        for idx, (sig_name, _, bt_res) in enumerate(results):
            alt  = (idx % 2 == 1)
            ms   = bt_res.metrics_strategy
            mb   = bt_res.metrics_benchmark

            excess_annual = ms["annual_return"] - mb["annual_return"]
            excess_sharpe = ms["sharpe"] - mb["sharpe"]

            values = [
                sig_name,
                self._pct(ms["total_return"]),
                self._pct(ms["annual_return"]),
                self._pct(ms["annual_vol"]),
                self._f4(ms["sharpe"]),
                self._pct(ms["max_drawdown"]),
                self._f4(ms["calmar"]),
                self._pct(ms["win_rate"]),
                ms.get("n_trades", 0),
                self._pct(ms.get("trade_win_rate")),
                self._pct(ms.get("holding_ratio")),
                self._f4(ms.get("odds_ratio")),
                self._pct(ms.get("excess_win_rate")),
                self._pct(mb["total_return"]),
                self._pct(mb["annual_return"]),
                self._f4(mb["sharpe"]),
                self._pct(mb["max_drawdown"]),
                self._pct(excess_annual),
                self._f4(excess_sharpe),
            ]
            self._write_data_row(ws, row=idx + 2, values=values, alt=alt)

            r = idx + 2
            # 策略总收益着色
            self._color_cell(ws.cell(r, 2),  ms["total_return"])
            # Sharpe 着色
            self._color_cell(ws.cell(r, 5),  ms["sharpe"])
            # 最大回撤着色（越小越好）
            self._color_cell(ws.cell(r, 6),  ms["max_drawdown"], reverse=True)
            # 交易胜率着色（>50% 为好）
            tw = ms.get("trade_win_rate")
            if tw is not None and not (isinstance(tw, float) and np.isnan(tw)):
                self._color_cell(ws.cell(r, 10), tw - 0.5)
            # 赔率着色（>1 为好）
            od = ms.get("odds_ratio")
            if od is not None and not (isinstance(od, float) and np.isnan(od)):
                self._color_cell(ws.cell(r, 12), od - 1.0)
            # 超额胜率着色（>50% 为好）
            ew = ms.get("excess_win_rate")
            if ew is not None and not (isinstance(ew, float) and np.isnan(ew)):
                self._color_cell(ws.cell(r, 13), ew - 0.5)
            # 超额年化收益着色
            self._color_cell(ws.cell(r, 18), excess_annual)

        self._auto_width(ws)

    # ── 统一入口 ─────────────────────────────────────────────────────────────

    def write(self,
              results: List[Tuple],
              asset_name: str,
              cfg: dict,
              save_path: str | Path) -> None:
        """
        生成完整 Excel 报告并保存。

        参数
        ----
        results   : [(sig_name, evaluator, BacktestResult), ...]
        asset_name: 资产显示名称
        cfg       : 完整配置字典
        save_path : 输出 Excel 文件路径
        """
        self._sheet_overview(results, asset_name, cfg)
        self._sheet_ic(results)
        self._sheet_signal(results)
        self._sheet_regression(results)
        self._sheet_robustness(results)
        self._sheet_backtest(
            results,
            benchmark_ratio=cfg.get("backtest", {}).get("benchmark_asset_ratio", 0.5),
        )

        save_path = Path(save_path)
        save_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(save_path)
        print(f"\n  ✓ Excel报告已保存：{save_path.resolve()}")


# ══════════════════════════════════════════════════════════════════════════════
# 多因子横向对比图（复用 example 中的逻辑）
# ══════════════════════════════════════════════════════════════════════════════

def plot_multi_factor_comparison(results: List[Tuple], save_path: Path) -> None:
    """生成四子图多因子横向对比图并保存。"""
    if len(results) < 2:
        return  # 单因子无需对比图

    names  = [r[0] for r in results]
    evs    = {r[0]: r[1] for r in results}
    scores = {r[0]: r[1].score() for r in results}
    n      = len(names)
    colors = plt.cm.tab10(np.linspace(0, 0.8, n))
    periods = [1, 5, 10, 20]

    fig, axes = plt.subplots(2, 2, figsize=(14, 10))
    fig.suptitle("择时信号横向对比", fontsize=14, fontweight="bold")

    # 子图A：各预测周期IC均值
    ax = axes[0, 0]
    for i, (name, ev) in enumerate(evs.items()):
        ic_means = [
            (ev._ic_results[p].ic_mean if ev._ic_results and p in ev._ic_results else 0)
            for p in periods
        ]
        ax.plot(periods, ic_means, marker="o", label=name,
                color=colors[i], linewidth=1.8, markersize=5)
    ax.axhline(0, color="black", linewidth=0.5, linestyle="--")
    ax.set_title("各预测周期IC均值", fontsize=11)
    ax.set_xlabel("预测周期（天）")
    ax.set_ylabel("IC均值")
    ax.legend(fontsize=9)
    ax.grid(True, alpha=0.3)

    # 子图B：1日ICIR
    ax = axes[0, 1]
    icirs = []
    for name in names:
        ev = evs[name]
        ic1 = (ev._ic_results or {}).get(1)
        icir = ic1.icir if ic1 and not np.isnan(ic1.icir) else 0.0
        icirs.append(icir)
    bars = ax.bar(range(n), icirs,
                  color=[colors[i] for i in range(n)], alpha=0.85, edgecolor="black")
    ax.axhline(0,   color="black", linewidth=0.5)
    ax.axhline(0.5, color="#43A047", linewidth=1.2, linestyle="--",
               alpha=0.8, label="ICIR=0.5（良好）")
    ax.axhline(-0.5, color="#E53935", linewidth=1.2, linestyle="--", alpha=0.8)
    ax.set_xticks(range(n))
    ax.set_xticklabels(names, fontsize=9, rotation=15)
    ax.set_title("ICIR（1日预测）", fontsize=11)
    ax.set_ylabel("ICIR")
    ax.legend(fontsize=8)

    # 子图C：三种方法胜率
    ax = axes[1, 0]
    methods = ["threshold", "moving_average", "percentile"]
    mlabels = ["阈值法", "均线法", "极值法"]
    x = np.arange(n)
    w = 0.25
    for j, (m, ml) in enumerate(zip(methods, mlabels)):
        wrs = []
        for name in names:
            ev = evs[name]
            res = (ev._signal_results or {}).get(m)
            wrs.append(res.full.overall_win_rate * 100 if res else 50.0)
        ax.bar(x + j * w, wrs, w, label=ml, alpha=0.85)
    ax.axhline(50, color="#E53935", linewidth=1.2, linestyle="--", label="50%基准")
    ax.set_xticks(x + w)
    ax.set_xticklabels(names, fontsize=9, rotation=15)
    ax.set_title("各方法胜率对比", fontsize=11)
    ax.set_ylabel("胜率（%）")
    ax.set_ylim(35, 75)
    ax.legend(fontsize=8)

    # 子图D：综合评分
    ax = axes[1, 1]
    composites = [scores[n].composite_score for n in names]
    bars = ax.bar(range(len(names)), composites,
                  color=[colors[i] for i in range(n)], alpha=0.85, edgecolor="black")
    ax.axhline(0.45, color="#43A047", linewidth=1.2, linestyle="--", label="B-级（0.45）")
    ax.axhline(0.75, color="#1565C0", linewidth=1.2, linestyle="--", label="A级（0.75）")
    ax.set_xticks(range(len(names)))
    ax.set_xticklabels(names, fontsize=9, rotation=15)
    ax.set_title("综合评分与评级", fontsize=11)
    ax.set_ylabel("评分 [0-1]")
    ax.set_ylim(0, 1.08)
    ax.legend(fontsize=8)
    for i, (bar, sc, name) in enumerate(zip(bars, composites, names)):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.012,
                f"{scores[name].grade}级",
                ha="center", va="bottom", fontsize=9, fontweight="bold")

    plt.tight_layout()
    save_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(save_path, dpi=100, bbox_inches="tight")
    plt.close(fig)
    print(f"  ✓ 多因子对比图：{save_path.name}")


# ══════════════════════════════════════════════════════════════════════════════
# 主程序
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    ROOT = Path(__file__).parent

    # ── 加载配置 ─────────────────────────────────────────────────────────────
    cfg = load_config(ROOT / "config.yaml") # -> cfg: dict

    asset_cfg  = cfg.get("asset",      {}) # -> asset_cfg: dict
    sig_cfg    = cfg.get("signals",    {}) # -> sig_cfg: dict
    eval_cfg   = cfg.get("evaluation", {}) # -> eval_cfg: dict
    bt_cfg     = cfg.get("backtest",   {}) # -> bt_cfg: dict
    out_cfg    = cfg.get("output",     {}) # -> out_cfg: dict

    asset_name  = asset_cfg.get("name",        "未知资产") # -> asset_name: str
    asset_code  = asset_cfg.get("code",        "") # -> asset_code: str
    price_field = asset_cfg.get("price_field", "CLOSE") # -> price_field: str

    workspace        = ROOT / out_cfg.get("workspace_dir", "workspace") # -> workspace: Path_str
    eval_subdir      = out_cfg.get("eval_plots_subdir",    "plots/eval") # -> eval_subdir: Path_str
    bt_subdir        = out_cfg.get("backtest_plots_subdir", "plots/backtest") # -> bt_subdir: Path_str
    excel_filename   = out_cfg.get("excel_filename",       "timing_report.xlsx") # -> excel_filename: Path_str
    dpi              = int(out_cfg.get("dpi", 100)) # -> dpi: int

    workspace.mkdir(parents=True, exist_ok=True) #创立文件夹

    print("=" * 68)
    print("  择时因子评价框架  |  主程序启动")
    print(f"  资产：{asset_name}（{asset_code}）")
    print("=" * 68)

    # ── 加载数据 ─────────────────────────────────────────────────────────────
    print("\n[Step 1] 加载数据...")
    data_cfg    = cfg.get("data", {}) # -> data_cfg: dict
    price_file  = data_cfg.get("price_file",  "price_df.csv") # -> price_file: Path_str
    signal_file = data_cfg.get("signal_file", "signal.csv") # -> signal_file: Path_str
    price_df, signal_df = load_data(ROOT / "data_ini", price_file, signal_file) # -> price_df: DataFrame, signal_df: DataFrame

    prices  = extract_price_series(price_df, asset_code, price_field) # -> prices: Series
    returns = prices.pct_change().dropna() # -> returns: Series

    # ── 日期范围过滤 ─────────────────────────────────────────────────────────
    start_date = data_cfg.get("start_date", "") or None # -> start_date: str or None
    end_date   = data_cfg.get("end_date",   "") or None # -> end_date: str or None
    if start_date or end_date: # 如果有开始或结束日期，则切片价格和信号表
        prices  = prices.loc[start_date:end_date]
        returns = returns.loc[start_date:end_date]
        sig_df_raw = signal_df.loc[start_date:end_date]
        date_info = f"{start_date or '起始'} → {end_date or '末尾'}"
        print(f"  日期过滤  ：{date_info}")
    else:
        sig_df_raw = signal_df

    sig_selected = select_signals(sig_df_raw, sig_cfg) #sig_selected: DataFrame 从信号表格中找到config中指定的信号
    signal_names = list(sig_selected.columns) # -> signal_names: list ->[str] 所有信号的名称

    print(f"  价格序列  ：{prices.index[0].date()} → {prices.index[-1].date()}"
          f"（共 {len(prices)} 条）")
    print(f"  选用信号  ：{signal_names}")

    # ── 逐信号评估与回测 ─────────────────────────────────────────────────────
    print("\n[Step 2] 逐信号评估与回测...")
    backtester = Backtester(bt_cfg) # -> backtester: Backtester对象
    all_results: List[Tuple] = []   # -> all_results: list ->[(sig_name, evaluator, BacktestResult)] 每个信号的回测结果

    for sig_name in signal_names:
        print(f"\n  ── {sig_name} {'─' * (52 - len(sig_name))}")
        raw_signal = sig_selected[sig_name].dropna() # -> raw_signal: Series 信号序列（删除缺失值）

        # ── 频率感知数据准备 ──────────────────────────────────────────
        # 自动检测信号频率，分离评估（信号自然频率）与回测（日度前向填充）
        eval_signal, eval_prices, eval_returns, eval_window, freq_label, median_gap = \
            _prepare_eval_data(raw_signal, prices,
                               int(eval_cfg.get("rolling_window", 252)))
        is_lowfreq = (median_gap > 2)
        if is_lowfreq:
            print(f"  信号频率  ：{freq_label}（检测间隔≈{median_gap}交易日）")
            print(f"  评估模式  ：降频期间收益（{len(eval_signal)}条）| 回测前向填充至日度")
            print(f"  滚动窗口  ：{int(eval_cfg.get('rolling_window', 252))}交易日 → 调整为{eval_window}期")
        else:
            print(f"  信号频率  ：{freq_label}")

        # 每个信号独立输出目录：workspace/{safe_name}/
        safe_name  = sig_name.replace("/", "_").replace(" ", "_").replace("\\", "_") # -> safe_name: str 安全的文件名
        sig_dir    = workspace / safe_name # -> sig_dir: Path_str 信号目录
        eval_dir   = sig_dir / eval_subdir # -> eval_dir: Path_str 评估目录
        bt_dir     = sig_dir / bt_subdir # -> bt_dir: Path_str 回测目录
        excel_path = sig_dir / excel_filename # -> excel_path: Path_str Excel文件路径
        eval_dir.mkdir(parents=True, exist_ok=True)
        bt_dir.mkdir(parents=True, exist_ok=True)

        # ── 因子评估（使用信号自然频率数据，避免重复计数）────────────
        ev = TimingFactorEvaluator(
            factor_name   = sig_name,
            forward_period= int(eval_cfg.get("forward_period", 1)),
            ic_method     = eval_cfg.get("ic_method", "pearson"),
        ) # -> ev: TimingFactorEvaluator对象
        ev.evaluate(
            factor                = eval_signal,
            returns               = eval_returns,
            prices                = eval_prices,
            preprocess            = bool(eval_cfg.get("preprocess", True)),
            rolling_window        = eval_window,
            run_robustness        = bool(eval_cfg.get("run_robustness", True)),
            run_rolling_regression= bool(eval_cfg.get("run_rolling_regression", False)),
            ic_periods            = eval_cfg.get("ic_periods", [1, 5, 10, 20]),
            signal_kwargs         = {
                **eval_cfg.get("signal_params", {}),
                "test_ratio": float(eval_cfg.get("signal_test_ratio", 0.30)),
            },
        ) # 对于一个信号进行打分，包含五个方面，IC评分，信号评分，回归评分和稳健性评分

        # 打印评估报告（控制台）
        ev.report()

        # 保存评估图表
        fig = ev.plot(figsize=(16, 13))
        eval_plot_path = eval_dir / f"{safe_name}_evaluation.png"
        fig.savefig(eval_plot_path, dpi=dpi, bbox_inches="tight")
        plt.close(fig)
        print(f"  ✓ 评估图：{eval_plot_path.relative_to(workspace)}")

        # ── 回测用信号：优先使用评分最高方法的离散信号 ──────────────
        # score() 在 report() 中已被调用，best_method/best_signal 已设置
        # 若未设置则回退到连续因子值
        _ = ev.score()  # 确保 best_method / best_signal 已填充
        best_signal_sparse = getattr(ev, "best_signal", None)

        if best_signal_sparse is not None:
            best_method_name = getattr(ev, "best_method", "unknown")
            print(f"  回测信号  ：最佳方法「{best_method_name}」的离散信号（+1/0/-1）")
            if is_lowfreq:
                factor_for_bt = (
                    best_signal_sparse.reindex(prices.index, method="ffill")
                    .fillna(0).astype(float)
                )
            else:
                factor_for_bt = best_signal_sparse.fillna(0).astype(float)
        else:
            # 回退：使用预处理后的连续因子值
            preprocessed_sparse = ev._factor if ev._factor is not None else eval_signal
            if is_lowfreq:
                factor_for_bt = preprocessed_sparse.reindex(prices.index, method="ffill")
            else:
                factor_for_bt = preprocessed_sparse
        bt_res = backtester.run(
            signal      = factor_for_bt,
            prices      = prices,
            signal_name = sig_name,
            save_dir    = bt_dir,
        ) # -> bt_res: BacktestResult对象 回测结果
        print(bt_res.summary())

        # ── 仓位表 ────────────────────────────────────────────────────────
        # 对齐信号值与仓位序列的共同索引
        common_idx = bt_res.positions.index
        pos_table = pd.DataFrame(
            {
                "信号值":       factor_for_bt.reindex(common_idx),
                "仓位":         bt_res.positions.map({0.0: "空仓", 1.0: "持仓"}),
                "仓位数值":     bt_res.positions,
                "策略日收益":   bt_res.strategy_returns,
                "基准日收益":   bt_res.benchmark_returns,
                "策略累计净值": (1 + bt_res.strategy_returns).cumprod(),
                "基准累计净值": (1 + bt_res.benchmark_returns).cumprod(),
            },
            index=common_idx,
        )
        pos_table.index.name = "日期"
        pos_path = sig_dir / "position_table.csv"
        pos_table.to_csv(pos_path, encoding="utf-8-sig")   # utf-8-sig 保证 Excel 直接打开不乱码
        print(f"  ✓ 仓位表：{pos_path.relative_to(workspace)}")

        # ── 单信号 Excel 报告 ──────────────────────────────────────────────
        ExcelReporter().write(
            results    = [(sig_name, ev, bt_res)],
            asset_name = asset_name,
            cfg        = cfg,
            save_path  = excel_path,
        )

        all_results.append((sig_name, ev, bt_res))

    # ── 多因子对比图（存放在 workspace 根目录）────────────────────────────────
    print("\n[Step 3] 生成多因子对比图...")
    plot_multi_factor_comparison(
        results   = all_results,
        save_path = workspace / "factor_comparison.png",
    )

    # ── 汇总输出 ─────────────────────────────────────────────────────────────
    print("\n" + "=" * 68)
    print("  全部完成！输出目录结构：")
    print(f"  {workspace.relative_to(ROOT)}/")
    for sig_name, ev, bt_res in all_results:
        safe = sig_name.replace("/", "_").replace(" ", "_").replace("\\", "_")
        print(f"  ├── {safe}/")
        print(f"  │   ├── {eval_subdir}/   （评估图）")
        print(f"  │   ├── {bt_subdir}/  （回测图）")
        print(f"  │   ├── position_table.csv  （仓位表）")
        print(f"  │   └── {excel_filename}")
    print(f"  └── factor_comparison.png   （多信号对比图）")
    print()
    print("  各信号综合评级：")
    for sig_name, ev, bt_res in all_results:
        sc = ev.score()
        ms = bt_res.metrics_strategy
        print(f"    {sig_name:<20}  评级={sc.grade}  "
              f"综合={sc.composite_score:.3f}  "
              f"年化={ms['annual_return']:+.2%}  "
              f"SR={ms['sharpe']:.3f}")
    print("=" * 68 + "\n")


if __name__ == "__main__":
    main()
